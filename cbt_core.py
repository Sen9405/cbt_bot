"""
🧠 КПТ-Терапия — ядро

Общая бизнес-логика для всех платформ (Telegram, MAX и др.)
"""
"""
🧠 **КПТ-Терапия Бот**
Telegram-бот для работы в рамках когнитивно-поведенческой терапии.

Разделы:
- 🧠 КПТ Дневник — запись ситуаций, мыслей, эмоций и реакций
- 📋 Планы на день — планирование задач с отметкой выполнения
"""
import sqlite3
import os
from datetime import datetime, date, timedelta
import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
import socket
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO
from dotenv import load_dotenv

# ========== КОНФИГУРАЦИЯ ==========
# Ищем .env сначала рядом с ботом, потом в текущей директории
_ENV_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), '.env')
if not os.path.exists(_ENV_PATH):
    _ENV_PATH = os.path.join(os.getcwd(), '.env')
load_dotenv(_ENV_PATH)
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_FILE = os.path.join(BASE_DIR, "cbt_data.db")

ADMIN_ID = int(os.getenv("ADMIN_TG_ID", "0"))  # заменить на свой Telegram ID в .env
MAX_FREE_RECORDS = 20
MAX_FREE_PLANS = 20
MAX_FREE_ACHIEVEMENTS = 20
DATE_FORMAT = "%d-%m-%Y %H:%M"
DATE_FORMAT_DISPLAY = "ДД-ММ-ГГГГ ЧЧ:ММ"
DATE_FORMAT_EXAMPLE = "23-04-2026 15:30"
DB_DATE_FORMAT = "%Y-%m-%d %H:%M"  # для хранения в БД стандартный

def fmt_dt(dt):
    """datetime → ДД-ММ-ГГГГ ЧЧ:ММ для показа"""
    return dt.strftime(DATE_FORMAT)

def fmt_date_str(date_str):
    """ГГГГ-ММ-ДД → ДД-ММ-ГГГГ для показа"""
    if not date_str:
        return ''
    if len(date_str) == 10 and date_str[:4].isdigit():
        return f"{date_str[8:10]}-{date_str[5:7]}-{date_str[:4]}"
    return date_str

def now_str():
    return datetime.now().strftime(DATE_FORMAT)

def db_to_display(db_date):
    """ГГГГ-ММ-ДД ЧЧ:ММ → ДД-ММ-ГГГГ ЧЧ:ММ"""
    if not db_date:
        return "-"
    try:
        dt = datetime.strptime(db_date, DB_DATE_FORMAT)
        return dt.strftime(DATE_FORMAT)
    except:
        return db_date

def display_to_db(display_date):
    """ДД-ММ-ГГГГ ЧЧ:ММ → ГГГГ-ММ-ДД ЧЧ:ММ"""
    dt = datetime.strptime(display_date, DATE_FORMAT)
    return dt.strftime(DB_DATE_FORMAT)

# ========== БАЗА ДАННЫХ ==========
def init_db():
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    
    # Создаём таблицу если её нет (простая и безопасная инициализация)
    c.execute('''CREATE TABLE IF NOT EXISTS cbt_records (
        id INTEGER,
        user_id INTEGER NOT NULL,
        created_at TEXT NOT NULL,
        situation TEXT,
        thought TEXT,
        emotion TEXT,
        body_reaction TEXT,
        behavior_reaction TEXT,
        confirmation TEXT,
        refutation TEXT,
        PRIMARY KEY (user_id, id)
    )''')
    
    # Проверяем, есть ли данные — если нет, таблица новая, всё ок
    c.execute("SELECT COUNT(*) FROM cbt_records")
    count = c.fetchone()[0]
    
    if count == 0:
        # Проверяем, не было ли старой таблицы с другими колонками
        c.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='cbt_records_old'")
        if c.fetchone():
            # Переносим данные из старой таблицы
            try:
                c.execute('SELECT * FROM cbt_records_old ORDER BY user_id, created_at ASC')
                old_records = c.fetchall()
                if old_records:
                    user_counters = {}
                    for rec in old_records:
                        uid = rec[1]
                        user_counters[uid] = user_counters.get(uid, 0) + 1
                        new_id = user_counters[uid]
                        c.execute('''INSERT OR IGNORE INTO cbt_records 
                            (id, user_id, created_at, situation, thought, emotion, body_reaction, behavior_reaction)
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?)''',
                            (new_id, uid, rec[2],
                             rec[3] if len(rec) > 3 else None,
                             rec[4] if len(rec) > 4 else None,
                             rec[5] if len(rec) > 5 else None,
                             rec[6] if len(rec) > 6 else None,
                             rec[7] if len(rec) > 7 else None))
                    print(f"✅ Перенесено {len(old_records)} старых записей")
                c.execute('DROP TABLE IF EXISTS cbt_records_old')
            except Exception as e:
                print(f"⚠️ Ошибка при переносе старых данных: {e}")
    
    # Миграция: добавляем колонки confirmation и refutation, если их нет
    for col in ['confirmation', 'refutation']:
        try:
            c.execute(f'ALTER TABLE cbt_records ADD COLUMN {col} TEXT')
        except:
            pass  # колонка уже существует
    
    c.execute('''CREATE TABLE IF NOT EXISTS daily_plans (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER NOT NULL,
        date TEXT NOT NULL,
        created_at TEXT NOT NULL
    )''')
    c.execute('''CREATE TABLE IF NOT EXISTS plan_items (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        plan_id INTEGER NOT NULL,
        text TEXT NOT NULL,
        done INTEGER DEFAULT 0,
        sort_order INTEGER DEFAULT 0,
        FOREIGN KEY (plan_id) REFERENCES daily_plans(id)
    )''')
    c.execute('''CREATE TABLE IF NOT EXISTS achievements (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER NOT NULL,
        text TEXT NOT NULL,
        created_at TEXT NOT NULL,
        position INTEGER DEFAULT 0
    )''')
    
    conn.commit()
    conn.close()

def save_record(user_id, data):
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    now = datetime.now().strftime(DB_DATE_FORMAT)
    
    # Следующий id для пользователя
    c.execute('SELECT COALESCE(MAX(id), 0) + 1 FROM cbt_records WHERE user_id = ?', (user_id,))
    next_id = c.fetchone()[0]
    
    c.execute('''
        INSERT INTO cbt_records (id, user_id, created_at, situation, thought, emotion, body_reaction, behavior_reaction, confirmation, refutation)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', (next_id, user_id, data.get('created_at', now), 
          data.get('situation'), data.get('thought'), 
          data.get('emotion'), data.get('body_reaction'), 
          data.get('behavior_reaction'), 
          data.get('confirmation'), data.get('refutation')))
    conn.commit()
    conn.close()
    return next_id

def get_records(user_id, days=7):
    start = (datetime.now() - timedelta(days=days)).strftime(DB_DATE_FORMAT[:10])
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute('''
        SELECT * FROM cbt_records 
        WHERE user_id = ? AND substr(created_at,1,10) >= ?
        ORDER BY created_at ASC
    ''', (user_id, start))
    records = c.fetchall()
    conn.close()
    return records

def get_records_by_period(user_id, start_date, end_date):
    start = start_date.strftime(DB_DATE_FORMAT[:10])
    end = end_date.strftime("%Y-%m-%d 23:59")
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute('''
        SELECT * FROM cbt_records 
        WHERE user_id = ? AND created_at >= ? AND created_at <= ?
        ORDER BY created_at ASC
    ''', (user_id, start, end))
    records = c.fetchall()
    conn.close()
    return records

def update_record(record_id, user_id, field, value):
    ALLOWED_FIELDS = {'created_at', 'situation', 'thought', 'emotion', 'body_reaction', 'behavior_reaction', 'confirmation', 'refutation'}
    if field not in ALLOWED_FIELDS:
        raise ValueError(f"Недопустимое поле: {field}")
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute(f'UPDATE cbt_records SET {field} = ? WHERE id = ? AND user_id = ?', (value, record_id, user_id))
    conn.commit()
    conn.close()

def get_record_by_id(record_id, user_id):
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute('SELECT * FROM cbt_records WHERE id = ? AND user_id = ?', (record_id, user_id))
    record = c.fetchone()
    conn.close()
    return record

def delete_record(record_id, user_id):
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute('DELETE FROM cbt_records WHERE id = ? AND user_id = ?', (record_id, user_id))
    conn.commit()
    conn.close()

def get_records_count(user_id):
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute('SELECT COUNT(*) FROM cbt_records WHERE user_id = ?', (user_id,))
    count = c.fetchone()[0]
    conn.close()
    return count

def get_or_create_plan(user_id, target_date=None):
    if target_date is None:
        target_date = date.today().strftime("%Y-%m-%d")
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute('SELECT id FROM daily_plans WHERE user_id = ? AND date = ?', (user_id, target_date))
    row = c.fetchone()
    if row:
        plan_id = row[0]
    else:
        now = datetime.now().strftime(DB_DATE_FORMAT)
        c.execute('INSERT INTO daily_plans (user_id, date, created_at) VALUES (?, ?, ?)',
                  (user_id, target_date, now))
        plan_id = c.lastrowid
    conn.commit()
    conn.close()
    return plan_id

def add_plan_item(plan_id, text):
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute('SELECT COALESCE(MAX(sort_order), -1) + 1 FROM plan_items WHERE plan_id = ?', (plan_id,))
    sort_order = c.fetchone()[0]
    c.execute('INSERT INTO plan_items (plan_id, text, done, sort_order) VALUES (?, ?, 0, ?)',
              (plan_id, text, sort_order))
    conn.commit()
    conn.close()

def get_plan_items(plan_id):
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute('SELECT id, text, done, sort_order FROM plan_items WHERE plan_id = ? ORDER BY sort_order', (plan_id,))
    items = c.fetchall()
    conn.close()
    return items

def toggle_plan_item(item_id):
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute('UPDATE plan_items SET done = CASE WHEN done = 0 THEN 1 ELSE 0 END WHERE id = ?', (item_id,))
    conn.commit()
    conn.close()

def delete_plan_item(item_id):
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute('DELETE FROM plan_items WHERE id = ?', (item_id,))
    conn.commit()
    conn.close()

def delete_plan(plan_id):
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    # Сначала удаляем все пункты
    c.execute('DELETE FROM plan_items WHERE plan_id = ?', (plan_id,))
    c.execute('DELETE FROM daily_plans WHERE id = ?', (plan_id,))
    conn.commit()
    conn.close()

def add_achievement(user_id, text):
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    c.execute('SELECT COALESCE(MAX(position), -1) + 1 FROM achievements WHERE user_id = ?', (user_id,))
    pos = c.fetchone()[0]
    c.execute('INSERT INTO achievements (user_id, text, created_at, position) VALUES (?, ?, ?, ?)',
              (user_id, text, now, pos))
    conn.commit()
    conn.close()

def get_user_achievements(user_id):
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute('SELECT id, text, position FROM achievements WHERE user_id = ? ORDER BY position', (user_id,))
    items = c.fetchall()
    conn.close()
    return items

def get_achievements_by_period(user_id, start_date, end_date):
    """Возвращает (position_1based, text, created_at) за период (от старых к новым)"""
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    start = start_date.strftime('%Y-%m-%d')
    end = end_date.strftime('%Y-%m-%d 23:59')
    c.execute('''
        SELECT id, text, created_at
        FROM achievements
        WHERE user_id = ? AND created_at >= ? AND created_at <= ?
        ORDER BY created_at ASC
    ''', (user_id, start, end))
    items = c.fetchall()
    conn.close()
    return [(i+1, item[1], item[2]) for i, item in enumerate(items)]

def get_achievements_all(user_id):
    """Возвращает (position_1based, text, created_at) для всех достижений"""
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute('''
        SELECT id, text, created_at
        FROM achievements
        WHERE user_id = ?
        ORDER BY position
    ''', (user_id,))
    items = c.fetchall()
    conn.close()
    return [(i+1, item[1], item[2]) for i, item in enumerate(items)]

def delete_achievement(item_id, user_id=None):
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    if user_id:
        c.execute('DELETE FROM achievements WHERE id = ? AND user_id = ?', (item_id, user_id))
    else:
        c.execute('DELETE FROM achievements WHERE id = ?', (item_id,))
    conn.commit()
    conn.close()

def update_achievement(item_id, new_text):
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute('UPDATE achievements SET text = ? WHERE id = ?', (new_text, item_id))
    conn.commit()
    conn.close()

def count_user_plans(user_id):
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute('SELECT COUNT(*) FROM daily_plans WHERE user_id = ?', (user_id,))
    count = c.fetchone()[0]
    conn.close()
    return count

def count_achievements(user_id):
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute('SELECT COUNT(*) FROM achievements WHERE user_id = ?', (user_id,))
    count = c.fetchone()[0]
    conn.close()
    return count

def clean_achievement_text(text):
    """Убирает нумерацию в начале"""
    import re
    text = text.strip()
    text = re.sub(r'^[\d]+[\.\)]\s*', '', text)
    text = re.sub(r'^[\d]+\.[\d]+[\.\)]\s*', '', text)
    text = re.sub(r'^[\d]+\s', '', text)
    return text.strip()

def get_user_plans(user_id, days=7):
    start = (date.today() - timedelta(days=days)).strftime("%Y-%m-%d")
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute('SELECT id, date FROM daily_plans WHERE user_id = ? AND date >= ? ORDER BY date DESC', (user_id, start))
    plans = c.fetchall()
    conn.close()
    return plans

def check_plan_exists(user_id, target_date):
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute('SELECT id FROM daily_plans WHERE user_id = ? AND date = ?', (user_id, target_date))
    row = c.fetchone()
    conn.close()
    if row:
        items = get_plan_items(row[0])
        return row[0], bool(items)
    return None, False

# ========== ОТОБРАЖЕНИЕ ==========
FIELD_NAMES = {
    'created_at': 'Дата и время',
    'situation': 'Ситуация',
    'thought': 'Мысль',
    'confirmation': 'Подтверждения мысли',
    'refutation': 'Опровержения мысли',
    'emotion': 'Эмоция',
    'body_reaction': 'Реакция → Тело',
    'behavior_reaction': 'Реакция → Поведение'
}

def format_record(record):
    if not record:
        return "❌ Запись не найдена"
    
    return (
        f"📋 **Запись #{record[0]}**\n"
        f"📅 {db_to_display(record[2])}\n"
        f"📍 **Ситуация:** {record[3] or '-'}\n"
        f"💭 **Мысль:** {record[4] or '-'}\n"
        f"✅ **Подтверждения мысли:** {record[8] or '-'}\n"
        f"❌ **Опровержения мысли:** {record[9] or '-'}\n"
        f"😌 **Эмоция:** {record[5] or '-'}\n"
        f"🔹 **Реакция - Тело:** {record[6] or '-'}\n"
        f"🔹 **Реакция - Поведение:** {record[7] or '-'}"
    )

# ========== ЭКСПОРТ В EXCEL ==========
def export_to_excel(records):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "КПТ Дневник"
    
    header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_alignment = Alignment(vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    headers = ['ID', 'Дата и время', 'Ситуация', 'Мысль', 'Подтверждения мысли', 'Опровержения мысли', 'Эмоция', 'Реакция - Тело', 'Реакция - Поведение']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    for row_idx, record in enumerate(records, 2):
        data = [record[0], db_to_display(record[2]), record[3] or '', record[4] or '', 
                record[8] or '', record[9] or '',
                record[5] or '', record[6] or '', record[7] or '']
        for col_idx, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = cell_alignment
            cell.border = thin_border
            cell.font = Font(name='Arial', size=10)
    
    widths = [6, 18, 40, 40, 40, 40, 30, 30, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w
    
    ws.row_dimensions[1].height = 30
    
    excel_data = BytesIO()
    wb.save(excel_data)
    excel_data.seek(0)
    return excel_data

def get_plans_by_period(user_id, start_date, end_date):
    """Возвращает [(date, plan_items_list), ...] за период"""
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    start = start_date.strftime('%Y-%m-%d')
    end = end_date.strftime('%Y-%m-%d 23:59')
    c.execute('''
        SELECT id, date, created_at
        FROM daily_plans
        WHERE user_id = ? AND date >= ? AND date <= ?
        ORDER BY date ASC
    ''', (user_id, start, end))
    plans = c.fetchall()
    conn.close()
    result = []
    for plan_id, plan_date, created_at in plans:
        items = get_plan_items(plan_id)
        result.append({
            'plan_id': plan_id,
            'date': plan_date,
            'created_at': created_at,
            'items': items  # [(id, text, done, sort_order)]
        })
    return result

def get_all_plans(user_id):
    """Возвращает все планы пользователя"""
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute('''
        SELECT id, date, created_at
        FROM daily_plans
        WHERE user_id = ?
        ORDER BY date ASC
    ''', (user_id,))
    plans = c.fetchall()
    conn.close()
    result = []
    for plan_id, plan_date, created_at in plans:
        items = get_plan_items(plan_id)
        result.append({
            'plan_id': plan_id,
            'date': plan_date,
            'created_at': created_at,
            'items': items
        })
    return result

def export_plans_to_excel(plans_data, period_label):
    """Создаёт Excel с планами. Столбцы: ID, Дата, Пункт плана, Выполнено, Дата создания"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Планы"
    
    header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_alignment = Alignment(vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    headers = ['ID', 'Дата', 'Пункт плана', 'Выполнено', 'Дата создания']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Цвета для чередования дат
    date_fill_1 = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    date_fill_2 = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    date_font = Font(name='Arial', bold=True, size=11, color='375623')
    
    row_idx = 2
    item_counter = 1
    color_toggle = True
    for plan in plans_data:
        # Формат даты ДД-ММ-ГГГГ
        raw_date = plan['date']
        try:
            dt = datetime.strptime(raw_date, "%Y-%m-%d")
            plan_date_display = dt.strftime("%d-%m-%Y")
        except:
            plan_date_display = raw_date
        
        items = plan['items']
        start_row = row_idx
        plan_bg = date_fill_1 if color_toggle else date_fill_2
        
        if not items:
            # Если нет пунктов, пишем одну строку "(нет пунктов)"
            data = ['', plan_date_display, '(нет пунктов)', '', '']
            for col_idx, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = cell_alignment
                cell.border = thin_border
                cell.font = Font(name='Arial', size=10)
            row_idx += 1
        else:
            for item in items:
                item_id, item_text, done, sort_order = item
                done_label = "✅" if done else "⬜"
                data = [item_counter, '', item_text or '', done_label,
                        db_to_display(plan['created_at']) if plan['created_at'] else '']
                for col_idx, value in enumerate(data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.alignment = cell_alignment
                    cell.border = thin_border
                    cell.font = Font(name='Arial', size=10)
                    cell.fill = plan_bg
                row_idx += 1
                item_counter += 1
            
            # Объединяем ячейки даты для всей группы
            if len(items) > 1:
                ws.merge_cells(start_row=start_row, start_column=2, end_row=start_row + len(items) - 1, end_column=2)
            # Пишем дату в первую ячейку группы
            date_cell = ws.cell(row=start_row, column=2, value=plan_date_display)
            date_cell.font = date_font
            date_cell.alignment = Alignment(vertical='center', horizontal='center', wrap_text=True)
            date_cell.border = thin_border
            date_cell.fill = plan_bg
        
        color_toggle = not color_toggle
    
    widths = [6, 14, 55, 12, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w
    
    ws.row_dimensions[1].height = 30
    
    # Авто-высота для строк с данными
    for r in range(2, row_idx):
        ws.row_dimensions[r].height = 24
    
    excel_data = BytesIO()
    wb.save(excel_data)
    excel_data.seek(0)
    return excel_data

def export_achievements_to_excel(records):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Достижения"
    
    header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_alignment = Alignment(vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    headers = ['ID', 'Достижение', 'Дата добавления']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    for row_idx, record in enumerate(records, 2):
        data = [record[0], record[1] or '', db_to_display(record[2]) if record[2] else '']
        for col_idx, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = cell_alignment
            cell.border = thin_border
            cell.font = Font(name='Arial', size=10)
    
    widths = [6, 60, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w
    
    ws.row_dimensions[1].height = 30
    
    excel_data = BytesIO()
    wb.save(excel_data)
    excel_data.seek(0)
    return excel_data


# При импорте не инициализируем БД автоматически
# Вызовите init_db() в своём боте
